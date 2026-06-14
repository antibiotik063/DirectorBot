from __future__ import annotations

import io
import logging
import re

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

from app.config import Settings

logger = logging.getLogger(__name__)

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
REPORT_WORKSHEET_TITLE = "Кибер 2.0"


def _normalize_title(value: str) -> str:
    text = value.strip().replace("ё", "е").replace("Ё", "Е")
    return re.sub(r"\s+", " ", text).casefold()


class GoogleSheetsService:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        credentials = Credentials.from_service_account_info(
            settings.service_account_info,
            scopes=SCOPES,
        )
        self.drive_service = build("drive", "v3", credentials=credentials, cache_discovery=False)

    def _download_workbook_bytes(self) -> io.BytesIO:
        if not self.settings.google_drive_file_id:
            raise RuntimeError("Не задан GOOGLE_DRIVE_FILE_ID")

        try:
            logger.info("Скачиваю XLSX из Google Drive по id %s", self.settings.google_drive_file_id)
            request = self.drive_service.files().get_media(fileId=self.settings.google_drive_file_id)
            buffer = io.BytesIO()
            downloader = MediaIoBaseDownload(buffer, request)

            done = False
            while not done:
                _, done = downloader.next_chunk()

            buffer.seek(0)
            return buffer
        except HttpError as exc:
            raise RuntimeError(f"Ошибка скачивания XLSX из Google Drive: {exc}") from exc

    def get_report_values(self) -> list[list[str]]:
        workbook_bytes = self._download_workbook_bytes()

        try:
            workbook = load_workbook(workbook_bytes, data_only=True, read_only=True)
        except Exception as exc:
            raise RuntimeError(f"Не удалось прочитать XLSX-файл: {exc}") from exc

        try:
            worksheet = next(
                (sheet for sheet in workbook.worksheets if _normalize_title(sheet.title) == _normalize_title(REPORT_WORKSHEET_TITLE)),
                None,
            )
            if worksheet is None:
                raise RuntimeError(f"Лист '{REPORT_WORKSHEET_TITLE}' не найден в XLSX-файле")

            values: list[list[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                values.append(["" if cell is None else str(cell).strip() for cell in row])

            logger.info("Прочитано %s строк из листа '%s'", len(values), worksheet.title)
            return values
        finally:
            workbook.close()
